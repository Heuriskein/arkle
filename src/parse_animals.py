import re
import json
from openpyxl import load_workbook

wb = load_workbook("arknovaanimals_VM_v2.xlsx")
ws = wb.active

VALID_CONTINENTS = {'Africa', 'Asia', 'America', 'Europe', 'Australia'}

TYPE_MAP = {
    'Predator':   'Predator',
    'Herbivore':  'Herbivore',
    'Bear':       'Bear',
    'Primate':    'Primate',
    'Reptile':    'Reptile',
    'Bird':       'Bird',
    'Sea Animal': 'Sea Animal',
    'Pet':        'Petting Zoo',
}

# Manual corrections: card_id -> dict of field overrides
CORRECTIONS = {
    439: {'name': 'Llama'},
    458: {'name': 'Japanese Macaque', 'continents': ['Asia']},
    467: {'name': 'Ecuadorian Squirrel Monkey'},
    536: {'name': 'Longhorn Cowfish'},
}

def title_case(s):
    return s.title() if s else ''

def parse_size(raw):
    if not raw:
        return 0
    m = re.match(r'^\((\d+)\)', raw)   # (N) Aq M  — sea animals
    if m:
        return int(m.group(1))
    m = re.match(r'^PZ\s*(\d+)', raw)  # PZ N      — petting zoo
    if m:
        return int(m.group(1))
    m = re.match(r'^(\d+)', raw)       # plain, NR, NW, etc.
    if m:
        return int(m.group(1))
    return 0

def parse_tags(type_raw, size_raw):
    tags = []
    for part in re.split(r'[/\n]', type_raw or ''):
        t = part.strip()
        t = re.sub(r'\s+x\d+\s*$', '', t)
        t = re.sub(r'\s+\d+\s*$', '', t).strip()
        if t in TYPE_MAP and TYPE_MAP[t] not in tags:
            tags.append(TYPE_MAP[t])
    size_token = size_raw.split()[0] if size_raw else ''
    if 'R' in size_token:
        tags.append('Rock')
    if 'W' in size_token:
        tags.append('Water')
    return tags

def parse_continents(raw):
    conts = []
    for part in (raw or '').split('\n'):
        c = part.strip()
        c = re.sub(r'\s+x\d+\s*$', '', c).strip()
        if c == 'Americas':
            c = 'America'
        if c in VALID_CONTINENTS and c not in conts:
            conts.append(c)
    return conts

animals = []

for row in ws.iter_rows(min_row=2, values_only=False):
    card_cell = row[0]  # Column A: Card #
    raw_id = card_cell.value
    if raw_id is None:
        continue
    try:
        card_id = int(raw_id)
    except (ValueError, TypeError):
        continue
    if not (401 <= card_id <= 560):
        continue

    def cell(col):  # col is 0-indexed
        v = row[col].value
        return str(v).strip() if v is not None else ''

    name      = title_case(cell(1))   # B: Animal Card Name
    size_raw  = cell(3)               # D: Enclosure size
    cost_raw  = cell(4)               # E: Cost
    type_raw  = cell(5)               # F: Type
    cont_raw  = cell(6)               # G: Continent
    bonus_raw = cell(9)               # J: Bonuses (A/C/R)

    size      = parse_size(size_raw)
    tags      = parse_tags(type_raw, size_raw)
    continents = parse_continents(cont_raw)
    cost      = int(cost_raw) if re.match(r'^\d+$', cost_raw) else 0
    appeal, conservation = 0, 0
    m = re.match(r'^(\d+)/(\d+)/', bonus_raw)
    if m:
        appeal       = int(m.group(1))
        conservation = int(m.group(2))

    fix = CORRECTIONS.get(card_id, {})
    if 'name' in fix:
        name = fix['name']
    if 'continents' in fix:
        continents = fix['continents']

    animals.append({
        'id': card_id,
        'name': name,
        'cost': cost,
        'size': size,
        'tags': tags,
        'continents': continents,
        'appeal': appeal,
        'conservation': conservation,
    })

# Emit animals.js
lines = ['const ANIMALS = [']
for a in animals:
    tags_js  = ', '.join(f'"{t}"' for t in a['tags'])
    conts_js = ', '.join(f'"{c}"' for c in a['continents'])
    lines.append(
        f'  {{ id:{a["id"]}, name:"{a["name"]}", cost:{a["cost"]}, '
        f'size:{a["size"]}, tags:[{tags_js}], continents:[{conts_js}], appeal:{a["appeal"]}, conservation:{a["conservation"]} }},'
    )
lines.append('];')

with open('animals.js', 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

print(f'Total: {len(animals)} animals\n')
spot = {401,406,416,426,432,451,458,469,494,519,529,551}
for a in animals:
    if a['id'] in spot:
        print(f"#{a['id']} {a['name']}: cost={a['cost']} size={a['size']} "
              f"tags={a['tags']} conts={a['continents']} appeal={a['appeal']}")
